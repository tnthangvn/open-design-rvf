#!/usr/bin/env python3
"""Generate template-roster.v1.xlsx for the closed-tournament (giải khép kín) import.
Spec: 1 flat sheet (Athletes) + _META + Huong_dan.
Header = 2 rows: row1 Vietnamese label (for Sở), row2 machine key (snake_case, the import contract).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

TEMPLATE_VERSION = "1.0"

# ── styles ────────────────────────────────────────────────────────────────
LABEL_FILL = PatternFill("solid", fgColor="1F4E78")   # dark blue
LABEL_FONT = Font(bold=True, color="FFFFFF", size=11)
KEY_FILL = PatternFill("solid", fgColor="D9E1F2")     # light blue
KEY_FONT = Font(italic=True, color="404040", size=9)
REQ_FILL = PatternFill("solid", fgColor="C00000")     # red for required label
META_HEAD = PatternFill("solid", fgColor="375623")    # green
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = Workbook()

# ── Sheet _META ───────────────────────────────────────────────────────────
meta = wb.active
meta.title = "_META"
meta_rows = [
    ("key", "value", "Mô tả"),
    ("template_version", TEMPLATE_VERSION, "Phiên bản template (KHÔNG sửa)"),
    ("tournament_code", "SO-HCM-2026-001", "Mã giải khép kín do Sở cấp"),
    ("sport_slug", "pickleball", "Slug bộ môn trong hệ thống"),
    ("format", "GROUP_KNOCKOUT", "KNOCKOUT | LEAGUE | GROUP_KNOCKOUT | SWISS"),
    ("gender", "MIXED", "MALE | FEMALE | MIXED | UNLIMITED"),
    ("age", "UNLIMITED", "UNLIMITED | U12 | U15 | U18 | U21 | AGE_18_PLUS ..."),
]
for r, row in enumerate(meta_rows, start=1):
    for c, val in enumerate(row, start=1):
        cell = meta.cell(row=r, column=c, value=val)
        cell.border = BORDER
        if r == 1:
            cell.fill = META_HEAD
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = CENTER
        else:
            cell.alignment = LEFT
            if c == 1:
                cell.font = Font(bold=True)
meta.column_dimensions["A"].width = 22
meta.column_dimensions["B"].width = 26
meta.column_dimensions["C"].width = 55

# ── Sheet Athletes ────────────────────────────────────────────────────────
ath = wb.create_sheet("Athletes")

# columns: (machine_key, vn_label, required, width)
COLS = [
    ("external_ref",       "Mã VĐV (của Sở)",        False, 16),
    ("team_ref",           "Mã đội/đôi",             True,  14),
    ("team_name",          "Tên đội/đôi",            False, 20),
    ("group_name",         "Bảng",                   False, 8),
    ("seed",               "Hạt giống",              False, 10),
    ("full_name",          "Họ tên VĐV",             True,  22),
    ("dob",                "Ngày sinh (YYYY-MM-DD)", True,  18),
    ("gender",             "Giới tính",              True,  12),
    ("email",              "Email",                  True,  26),
    ("phone",              "SĐT",                    False, 14),
    ("is_captain",         "Đội trưởng?",            False, 11),
    ("guardian_name",      "GH - Họ tên",            False, 20),
    ("guardian_relation",  "GH - Quan hệ",           False, 12),
    ("guardian_phone",     "GH - SĐT",               False, 14),
    ("guardian_email",     "GH - Email",             False, 24),
    ("guardian_id_number", "GH - CCCD/CMND",         False, 16),
    ("guardian_address",   "GH - Địa chỉ",           False, 30),
]

# Row 1: VN labels  |  Row 2: machine keys
for c, (key, label, required, width) in enumerate(COLS, start=1):
    col_letter = ath.cell(row=1, column=c).column_letter
    ath.column_dimensions[col_letter].width = width

    lab = ath.cell(row=1, column=c, value=(label + (" *" if required else "")))
    lab.fill = REQ_FILL if required else LABEL_FILL
    lab.font = LABEL_FONT
    lab.alignment = CENTER
    lab.border = BORDER

    k = ath.cell(row=2, column=c, value=key)
    k.fill = KEY_FILL
    k.font = KEY_FONT
    k.alignment = CENTER
    k.border = BORDER

# Sample rows (start row 3): D01 = đôi người lớn; D02 = đôi vị thành niên kèm giám hộ
samples = [
    # ext, team_ref, team_name, group, seed, name, dob, gender, email, phone, captain, g_name, g_rel, g_phone, g_email, g_id, g_addr
    ("", "D01", "Đội Rồng Vàng", "A", 1, "Nguyễn Văn An", "1998-05-10", "male",   "an.nguyen@example.com",  "0901234567", "TRUE",  "", "", "", "", "", ""),
    ("", "D01", "Đội Rồng Vàng", "A", 1, "Trần Thị Bình", "2000-03-22", "female", "binh.tran@example.com",  "",           "FALSE", "", "", "", "", "", ""),
    ("", "D02", "Đội Sao Mai",   "B", 2, "Lê Minh Khôi",  "2012-09-01", "male",   "khoi.le@example.com",    "",           "TRUE",  "Lê Văn Cường", "Cha", "0912345678", "cuong.le@example.com", "012345678901", "123 Lê Lợi, Q1, TP.HCM"),
    ("", "D02", "Đội Sao Mai",   "B", 2, "Phạm Gia Hân",  "2013-07-15", "female", "han.pham@example.com",   "",           "FALSE", "Phạm Thị Mai", "Mẹ", "0987654321", "mai.pham@example.com", "079876543210", "45 Hai Bà Trưng, Q1, TP.HCM"),
]
for r, row in enumerate(samples, start=3):
    for c, val in enumerate(row, start=1):
        cell = ath.cell(row=r, column=c, value=val)
        cell.border = BORDER
        cell.alignment = LEFT

ath.freeze_panes = "A3"  # keep both header rows visible

# Data validations (dropdowns) applied generously over data area
dv_gender = DataValidation(type="list", formula1='"male,female,other"', allow_blank=False)
dv_captain = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
ath.add_data_validation(dv_gender)
ath.add_data_validation(dv_captain)
gender_col = COLS_index = [c for c, col in enumerate(COLS, 1) if col[0] == "gender"][0]
captain_col = [c for c, col in enumerate(COLS, 1) if col[0] == "is_captain"][0]
g_letter = ath.cell(row=1, column=gender_col).column_letter
c_letter = ath.cell(row=1, column=captain_col).column_letter
dv_gender.add(f"{g_letter}3:{g_letter}1000")
dv_captain.add(f"{c_letter}3:{c_letter}1000")

# ── Sheet Huong_dan (data dictionary) ─────────────────────────────────────
hd = wb.create_sheet("Huong_dan")
hd_header = ("Cột (key máy)", "Nhãn", "Bắt buộc", "Mô tả / Quy tắc", "Ví dụ")
dict_rows = [
    ("external_ref", "Mã VĐV (của Sở)", "Không", "Mã định danh ổn định của Sở. Nếu có → dùng làm khoá chống trùng ưu tiên số 1.", "VDV-00123"),
    ("team_ref", "Mã đội/đôi", "CÓ", "Khoá gom nhóm. Các VĐV cùng 1 đội/đôi phải có cùng team_ref.", "D01"),
    ("team_name", "Tên đội/đôi", "Không", "Tên hiển thị. Giống nhau trong cùng team_ref.", "Đội Rồng Vàng"),
    ("group_name", "Bảng", "Không", "Phân bảng sẵn (format GROUP_KNOCKOUT). Để trống nếu không chia bảng.", "A"),
    ("seed", "Hạt giống", "Không", "Số nguyên xếp hạt giống của đội.", "1"),
    ("full_name", "Họ tên VĐV", "CÓ", "Họ tên đầy đủ.", "Nguyễn Văn An"),
    ("dob", "Ngày sinh", "CÓ", "Định dạng YYYY-MM-DD. Dùng để kiểm tra độ tuổi & yêu cầu giám hộ.", "2012-09-01"),
    ("gender", "Giới tính", "CÓ", "male | female | other.", "male"),
    ("email", "Email", "CÓ", "Khoá định danh chính của VĐV. Mỗi VĐV nên có email riêng (giám hộ tạo giúp nếu nhỏ).", "an.nguyen@example.com"),
    ("phone", "SĐT", "Không", "Chỉ là thông tin, KHÔNG dùng làm khoá.", "0901234567"),
    ("is_captain", "Đội trưởng?", "Không", "TRUE/FALSE. Mỗi team_ref đúng 1 dòng TRUE.", "TRUE"),
    ("guardian_name", "GH - Họ tên", "Bắt buộc nếu <18", "Họ tên người giám hộ.", "Lê Văn Cường"),
    ("guardian_relation", "GH - Quan hệ", "Bắt buộc nếu <18", "Quan hệ với VĐV.", "Cha"),
    ("guardian_phone", "GH - SĐT", "Bắt buộc nếu <18", "SĐT giám hộ - kênh liên lạc cho VĐV nhỏ tuổi.", "0912345678"),
    ("guardian_email", "GH - Email", "Bắt buộc nếu <18", "Email giám hộ - nhận OTP/kích hoạt thay VĐV nhỏ tuổi.", "cuong.le@example.com"),
    ("guardian_id_number", "GH - CCCD/CMND", "Không", "Số CCCD/CMND người giám hộ.", "012345678901"),
    ("guardian_address", "GH - Địa chỉ", "Không", "Địa chỉ người giám hộ.", "123 Lê Lợi, Q1, TP.HCM"),
]
notes = [
    ("", "", "", "", ""),
    ("LƯU Ý CHUNG", "", "", "", ""),
    ("1", "", "", "Mỗi dòng = 1 vận động viên. Đội/đôi được gom theo cột team_ref.", ""),
    ("2", "", "", "KHÔNG xoá/đổi tên dòng key máy (dòng 2 của sheet Athletes) - đó là hợp đồng để hệ thống đọc.", ""),
    ("3", "", "", "Có thể thêm cột mới ở cuối; cột không nhận diện sẽ được bỏ qua khi import.", ""),
    ("4", "", "", "VĐV < 18 tuổi (tính theo dob) bắt buộc có đủ thông tin người giám hộ (GH).", ""),
    ("5", "", "", "Chống trùng khi import lại: external_ref → email → (full_name+dob+gender, gắn cờ review).", ""),
]
for r, row in enumerate([hd_header] + dict_rows + notes, start=1):
    for c, val in enumerate(row, start=1):
        cell = hd.cell(row=r, column=c, value=val)
        cell.border = BORDER if r == 1 or r <= 1 + len(dict_rows) else None
        if r == 1:
            cell.fill = LABEL_FILL
            cell.font = LABEL_FONT
            cell.alignment = CENTER
        else:
            cell.alignment = LEFT
            if c == 1:
                cell.font = Font(bold=True)
for col, w in zip("ABCDE", (20, 18, 16, 60, 26)):
    hd.column_dimensions[col].width = w
hd.freeze_panes = "A2"

# move _META first, Athletes second, Huong_dan third
wb.move_sheet("Athletes", offset=-1)

out = "/var/www/free-time/rvf-be/zandbox/templates/template-roster.v1.xlsx"
import os
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print("Saved:", out)
